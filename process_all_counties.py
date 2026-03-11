import pandas as pd
import openpyxl
import re
import os

def clean_name(name):
    if pd.isna(name): return ''
    return re.sub(r'[^a-z]', '', str(name).lower())

def process_all_counties(officials_file, voters_file, tracker_file):
    print("Loading data...")
    try:
        officials = pd.read_excel(officials_file)
        repubs = pd.read_csv(voters_file, dtype=str)
    except Exception as e:
        print(f"Error loading files: {e}")
        return

    # Clean repubs data once
    print("Cleaning voter names...")
    repubs['clean_first'] = repubs['FirstName'].apply(clean_name)
    repubs['clean_last'] = repubs['LastName'].apply(clean_name)
    repubs['clean_city'] = repubs['PrimaryCity'].apply(clean_name)

    # Get list of valid counties to process
    valid_counties = []
    for c in officials['County'].dropna().unique():
        if str(c).lower().strip() not in ['worcester', 'norfolk', 'key resources for additional research']:
            valid_counties.append(str(c).title())
    
    # Preload workbook
    try:
        wb = openpyxl.load_workbook(tracker_file)
        if 'Worcester County' not in wb.sheetnames:
            print("Error: Could not find 'Worcester County' template sheet!")
            return
    except Exception as e:
        print(f"Error loading tracker: {e}")
        return

    for county_name in valid_counties:
        print(f"\nProcessing {county_name} County...")
        county_officials = officials[officials['County'].str.lower() == county_name.lower()].copy()
        print(f"  Total officials: {len(county_officials)}")
        
        matched_rows = []
        for idx, row in county_officials.iterrows():
            name_str = str(row['Name'])
            name_parts = name_str.lower().split()
            if len(name_parts) < 2:
                continue
            
            first_name_candidate = clean_name(name_parts[0])
            last_name_candidate = clean_name(name_parts[-1])
            city_candidate = clean_name(row['Municipality'])
            
            # Exact Match
            match_cond = (
                (repubs['clean_first'] == first_name_candidate) & 
                (repubs['clean_last'] == last_name_candidate) &
                (repubs['clean_city'] == city_candidate)
            )
            
            matches_found = repubs[match_cond]
            if len(matches_found) > 0:
                match = matches_found.iloc[0]
                matched_rows.append({
                    'First': match.get('FirstName', ''),
                    'MI': match.get('MiddleName', ''), 
                    'Last': match.get('LastName', ''),
                    'Suffix': match.get('SuffixName', ''),
                    'Address': match.get('PrimaryAddress1', ''),
                    'Municipality': match.get('PrimaryCity', ''),
                    'Email': '', 
                    'Phone': match.get('PrimaryPhone', ''),
                    'Type': row.get('Committee', '')
                })

        print(f"  Found {len(matched_rows)} exact matches.")
        
        target_sheet_name = f'{county_name} County'
        if target_sheet_name in wb.sheetnames:
            print(f"  Replacing existing {target_sheet_name} sheet.")
            del wb[target_sheet_name]

        source_sheet = wb['Worcester County']
        target_sheet = wb.copy_worksheet(source_sheet)
        target_sheet.title = target_sheet_name

        # Clear data from columns 1 to 9 starting from row 4
        for row in range(4, target_sheet.max_row + 1):
            for col in range(1, 10):
                target_sheet.cell(row=row, column=col).value = None

        # Insert matches
        for row_idx, data_row in enumerate(matched_rows, start=4):
            target_sheet.cell(row=row_idx, column=1).value = data_row['First']
            target_sheet.cell(row=row_idx, column=2).value = data_row['MI']
            target_sheet.cell(row=row_idx, column=3).value = data_row['Last']
            target_sheet.cell(row=row_idx, column=4).value = data_row['Suffix']
            target_sheet.cell(row=row_idx, column=5).value = data_row['Address']
            target_sheet.cell(row=row_idx, column=6).value = data_row['Municipality']
            target_sheet.cell(row=row_idx, column=7).value = data_row['Email']
            target_sheet.cell(row=row_idx, column=8).value = data_row['Phone']
            target_sheet.cell(row=row_idx, column=9).value = data_row['Type']

        # Update right-side municipalities
        county_munis = sorted(county_officials['Municipality'].dropna().unique())
        
        for row in range(4, target_sheet.max_row + 1):
            for col in range(10, 15):
                target_sheet.cell(row=row, column=col).value = None

        for row_idx, muni in enumerate(county_munis, start=4):
            target_sheet.cell(row=row_idx, column=10).value = muni

    print("\nSaving tracker...")
    wb.save(tracker_file)
    print("Done!")

if __name__ == "__main__":
    officials_file = r'c:\Users\tbroo\Desktop\Coding\MA_Municipal_Officials.xlsx'
    voters_file = r'c:\Users\tbroo\Desktop\Coding\MyExport_3627.csv'
    tracker_file = r'c:\Users\tbroo\Desktop\Coding\2026 Ex-Officio Delegate Tracker.xlsx'
    process_all_counties(officials_file, voters_file, tracker_file)
