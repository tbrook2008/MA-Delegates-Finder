import pandas as pd
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
import argparse
import re
import os

def clean_name(name):
    if pd.isna(name): return ''
    return re.sub(r'[^a-z]', '', str(name).lower())

def process_county(county_name, officials_file, voters_file, tracker_file):
    print(f"Processing {county_name} County...")

    try:
        officials = pd.read_excel(officials_file)
        county_officials = officials[officials['County'].str.lower() == county_name.lower()].copy()
        print(f'Total officials in {county_name} County: {len(county_officials)}')
    except Exception as e:
        print(f'Error reading officials: {e}')
        return

    try:
        repubs = pd.read_csv(voters_file, dtype=str)
        print(f'Total republicans in CSV: {len(repubs)}')
    except Exception as e:
        print(f'Error reading republicans CSV: {e}')
        return

    repubs['clean_first'] = repubs['FirstName'].apply(clean_name)
    repubs['clean_last'] = repubs['LastName'].apply(clean_name)
    repubs['clean_city'] = repubs['PrimaryCity'].apply(clean_name)

    matched_rows = []

    for idx, row in county_officials.iterrows():
        name_str = str(row['Name'])
        name_parts = name_str.lower().split()
        if len(name_parts) < 2:
            continue
        
        first_name_candidate = clean_name(name_parts[0])
        last_name_candidate = clean_name(name_parts[-1])
        city_candidate = clean_name(row['Municipality'])
        
        # Exact Match: First Name, Last Name, and City
        match_cond = (
            (repubs['clean_first'] == first_name_candidate) & 
            (repubs['clean_last'] == last_name_candidate) &
            (repubs['clean_city'] == city_candidate)
        )
        
        matches_found = repubs[match_cond]
        if len(matches_found) > 0:
            match = matches_found.iloc[0]
            matched_rows.append({
                'First': match.get('FirstName', ''),
                'MI': match.get('MiddleName', ''), 
                'Last': match.get('LastName', ''),
                'Suffix': match.get('SuffixName', ''),
                'Address': match.get('PrimaryAddress1', ''),
                'Municipality': match.get('PrimaryCity', ''),
                'Email': '', 
                'Phone': match.get('PrimaryPhone', ''),
                'Type': row.get('Committee', '')
            })

    print(f"Found {len(matched_rows)} exact matches.")
    
    # Open the existing Excel file and make a copy of the Worcester County sheet
    wb = openpyxl.load_workbook(tracker_file)
    target_sheet_name = f'{county_name} County'
    
    if target_sheet_name in wb.sheetnames:
        print(f"{target_sheet_name} sheet already exists. Removing it to recreate.")
        del wb[target_sheet_name]

    if 'Worcester County' not in wb.sheetnames:
        print("Error: Could not find 'Worcester County' template sheet!")
        return

    source_sheet = wb['Worcester County']
    target_sheet = wb.copy_worksheet(source_sheet)
    target_sheet.title = target_sheet_name

    # Clear the data in the left table
    for row in range(4, target_sheet.max_row + 1):
        for col in range(1, 10):
            target_sheet.cell(row=row, column=col).value = None

    # Insert the new data
    for row_idx, data_row in enumerate(matched_rows, start=4):
        target_sheet.cell(row=row_idx, column=1).value = data_row['First']
        target_sheet.cell(row=row_idx, column=2).value = data_row['MI']
        target_sheet.cell(row=row_idx, column=3).value = data_row['Last']
        target_sheet.cell(row=row_idx, column=4).value = data_row['Suffix']
        target_sheet.cell(row=row_idx, column=5).value = data_row['Address']
        target_sheet.cell(row=row_idx, column=6).value = data_row['Municipality']
        target_sheet.cell(row=row_idx, column=7).value = data_row['Email']
        target_sheet.cell(row=row_idx, column=8).value = data_row['Phone']
        target_sheet.cell(row=row_idx, column=9).value = data_row['Type']

    # Update the right-side table with municipalities of this county
    county_munis = sorted(county_officials['Municipality'].dropna().unique())

    # Clear right side table
    for row in range(4, target_sheet.max_row + 1):
        for col in range(10, 15):
            target_sheet.cell(row=row, column=col).value = None

    for row_idx, muni in enumerate(county_munis, start=4):
        target_sheet.cell(row=row_idx, column=10).value = muni
        
    wb.save(tracker_file)
    print(f"Saved {target_sheet_name} to {tracker_file}")

if __name__ == "__main__":
    parser = argparse.ArgumentParser(description='Cross reference county officials with voters')
    parser.add_argument('county_name', help='Name of the county to process (e.g. "Norfolk", "Plymouth")')
    parser.add_argument('voters_file', help='Path to the CSV file with Republican voters for this county/counties')
    parser.add_argument('--officials_file', default='MA_Municipal_Officials.xlsx')
    parser.add_argument('--tracker_file', default='2026 Ex-Officio Delegate Tracker.xlsx')
    
    args = parser.parse_args()
    process_county(args.county_name, args.officials_file, args.voters_file, args.tracker_file)
