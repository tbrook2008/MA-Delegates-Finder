import pandas as pd
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
import shutil
import re
import os

excel_path = '2026 Ex-Officio Delegate Tracker.xlsx'

# Load the exact matches found previously
# We will just do the match again quickly to get all fields from repubs
officials = pd.read_excel('MA_Municipal_Officials.xlsx')
norfolk_officials = officials[officials['County'].str.lower() == 'norfolk'].copy()

repubs = pd.read_csv('MyExport_2820.csv', dtype=str)

def clean_name(name):
    if pd.isna(name): return ''
    return re.sub(r'[^a-z]', '', str(name).lower())

repubs['clean_first'] = repubs['FirstName'].apply(clean_name)
repubs['clean_last'] = repubs['LastName'].apply(clean_name)
repubs['clean_city'] = repubs['PrimaryCity'].apply(clean_name)

matched_rows = []

for idx, row in norfolk_officials.iterrows():
    name_str = str(row['Name'])
    name_parts = name_str.lower().split()
    if len(name_parts) < 2:
        continue
    
    first_name_candidate = clean_name(name_parts[0])
    last_name_candidate = clean_name(name_parts[-1])
    city_candidate = clean_name(row['Municipality'])
    
    # Exact Match: First Name, Last Name, and City
    match_cond = (
        (repubs['clean_first'] == first_name_candidate) & 
        (repubs['clean_last'] == last_name_candidate) &
        (repubs['clean_city'] == city_candidate)
    )
    
    matches_found = repubs[match_cond]
    if len(matches_found) > 0:
        match = matches_found.iloc[0]
        # First, MI, Last, Suffix, Address, Municipality, Email, Phone, Type
        # "Type" seems to be Committee from officials
        matched_rows.append({
            'First': match.get('FirstName', ''),
            'MI': match.get('MiddleName', ''), # might not exist, we'll try to extract from name_parts if needed, or leave blank
            'Last': match.get('LastName', ''),
            'Suffix': match.get('SuffixName', ''),
            'Address': match.get('PrimaryAddress1', ''),
            'Municipality': match.get('PrimaryCity', ''),
            'Email': '', # Not in voter file
            'Phone': match.get('PrimaryPhone', ''),
            'Type': row.get('Committee', '')
        })

print(f"Found {len(matched_rows)} exact matches.")
matches_df = pd.DataFrame(matched_rows)

# Now, open the existing Excel file and make a copy of the Worcester County sheet
wb = openpyxl.load_workbook(excel_path)
if 'Norfolk County' in wb.sheetnames:
    print("Norfolk County sheet already exists. Removing it to recreate.")
    del wb['Norfolk County']

source_sheet = wb['Worcester County']
target_sheet = wb.copy_worksheet(source_sheet)
target_sheet.title = 'Norfolk County'

# Clear the data in the left table (columns 1 to 9 starting from row 4/5)
# In Worcester County, headers are actually at row 3 (excel row 3 is index 3)
# Row 3 contains 'First', 'MI', 'Last', etc.
# Data starts at row 4.
for row in range(4, target_sheet.max_row + 1):
    for col in range(1, 10):
        target_sheet.cell(row=row, column=col).value = None

# Insert the new data
for row_idx, data_row in enumerate(matched_rows, start=4):
    target_sheet.cell(row=row_idx, column=1).value = data_row['First']
    target_sheet.cell(row=row_idx, column=2).value = data_row['MI']
    target_sheet.cell(row=row_idx, column=3).value = data_row['Last']
    target_sheet.cell(row=row_idx, column=4).value = data_row['Suffix']
    target_sheet.cell(row=row_idx, column=5).value = data_row['Address']
    target_sheet.cell(row=row_idx, column=6).value = data_row['Municipality']
    target_sheet.cell(row=row_idx, column=7).value = data_row['Email']
    target_sheet.cell(row=row_idx, column=8).value = data_row['Phone']
    target_sheet.cell(row=row_idx, column=9).value = data_row['Type']

# Update the side-table with Norfolk municipalities instead of Worcester?
# Let's get the unique list of Norfolk municipalities to replace the right side
norfolk_munis = sorted(norfolk_officials['Municipality'].dropna().unique())

# The right side table starts at column 10 (J) row 4. Header is at row 3 (J3: 'Municipalities')
# Let's clear it first
for row in range(4, target_sheet.max_row + 1):
    for col in range(10, 15):
        target_sheet.cell(row=row, column=col).value = None

for row_idx, muni in enumerate(norfolk_munis, start=4):
    target_sheet.cell(row=row_idx, column=10).value = muni
    
wb.save(excel_path)
print("Saved to 2026 Ex-Officio Delegate Tracker.xlsx")
